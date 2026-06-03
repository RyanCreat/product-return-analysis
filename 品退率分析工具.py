#!/usr/bin/env python3
"""
品退率横向对比分析工具
自动识别店铺，生成横向对比汇总表
适用平台：Windows / macOS
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import json
import threading
from pathlib import Path

# ====================== 配置文件 ======================
CONFIG_FILE = Path(__file__).parent / "品退率分析工具_config.json"

DEFAULT_PERIOD_KEYWORDS = {
    "7天": ["7天", "近7天"],
    "30天": ["30天", "近30天"],
}


def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ====================== 核心逻辑 ======================
def safe_int(v):
    if v is None or v == "" or v == "-":
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def safe_float(v):
    if v is None or v == "" or v == "-":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def extract_store_name(filename):
    """从文件名中自动提取店铺名称。
    去掉「-7天」「近7天」「-30天」「近30天」及其后面的内容，
    剩下的就是店铺名称。
    """
    base = filename.replace(".xlsx", "").replace(".xls", "")
    # 尝试各种分隔模式，取最短的那个作为店铺名
    patterns = [
        r"[-_]?近?7天.*$",
        r"[-_]?近?30天.*$",
        r"[-_]7天.*$",
        r"[-_]30天.*$",
    ]
    best = base
    for pat in patterns:
        m = re.search(pat, base)
        if m:
            candidate = base[:m.start()]
            if len(candidate) < len(best):
                best = candidate
    return best.strip()


def match_period(filename):
    """根据文件名识别时间周期"""
    for period, keywords in DEFAULT_PERIOD_KEYWORDS.items():
        for kw in keywords:
            if kw in filename:
                return period
    return None


def scan_folder(folder_path):
    """扫描文件夹，自动识别店铺和周期"""
    matches = {}  # {store: {period: filepath}}
    unmatched_store = []  # 无法识别周期的文件
    skipped = []  # 非xlsx文件

    for fname in os.listdir(folder_path):
        if not fname.endswith(".xlsx"):
            continue

        filepath = os.path.join(folder_path, fname)
        store = extract_store_name(fname)
        period = match_period(fname)

        if period is None:
            unmatched_store.append(fname)
            continue

        if store not in matches:
            matches[store] = {}
        matches[store][period] = filepath

    return matches, unmatched_store


def load_all_data(matches):
    """加载所有文件数据"""
    data = {}
    for store, periods in matches.items():
        data[store] = {}
        for period, filepath in periods.items():
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            period_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cat1, cat2, cat3, qty, rate, weight, industry = row
                if cat3:
                    period_data[str(cat3)] = {
                        "qty": safe_int(qty),
                        "rate": safe_float(rate),
                        "weight": safe_float(weight) or 0,
                        "industry": str(industry) if industry else "",
                    }
            data[store][period] = period_data
    return data


def calc_internal_avg(data, stores, cat, period):
    total_qty = 0
    weighted_sum = 0
    for store in stores:
        d = data[store].get(period, {}).get(cat, {})
        qty = d.get("qty", 0)
        rate = d.get("rate")
        if qty > 0 and rate is not None:
            total_qty += qty
            weighted_sum += rate * qty
    if total_qty > 0:
        return weighted_sum / total_qty, total_qty
    return None, 0


def has_sales(d):
    return d.get("qty", 0) > 0 and d.get("rate") is not None


# ====================== 样式常量 ======================
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_HEADER_FONT = Font(name="微软雅黑", bold=True, size=9)
CAT_FONT = Font(name="微软雅黑", bold=True, size=10)
DATA_FONT = Font(name="微软雅黑", size=10)
AVG_FONT = Font(name="微软雅黑", bold=True, size=10, color="2F5496")
NA_FONT = Font(name="微软雅黑", size=9, color="AAAAAA")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
AVG_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
NA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def industry_fill(level, has_sales_flag):
    if not has_sales_flag:
        return NA_FILL
    s = str(level)
    if "优于" in s: return GREEN_FILL
    if "1-2倍" in s: return YELLOW_FILL
    if "2-3倍" in s: return ORANGE_FILL
    if "3-5倍" in s or "5倍以上" in s: return RED_FILL
    return WHITE_FILL


def industry_label(level, has_sales_flag):
    if not has_sales_flag:
        return "无销售"
    s = str(level)
    if "优于" in s: return "达标"
    if "1-2倍" in s: return "轻微超标"
    if "2-3倍" in s: return "超标"
    if "3-5倍" in s or "5倍以上" in s: return "严重超标"
    return "-"


def label_color(label):
    if label == "达标": return "008000"
    if label == "轻微超标": return "806000"
    if label == "超标": return "C00000"
    if label == "严重超标": return "C00000"
    if label == "无销售": return "AAAAAA"
    return "000000"


def generate_excel(data, stores, output_path, selected_periods, progress_callback=None):
    """
    selected_periods: list, e.g. ["7天", "30天"] or ["7天"] or ["30天"]
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # collect categories per period
    all_cats = {}
    for period in selected_periods:
        cats = set()
        for store in stores:
            if period in data[store]:
                cats.update(data[store][period].keys())
        all_cats[period] = cats

    def sort_cats(cats, period):
        cat_vol = {}
        for c in cats:
            _, total = calc_internal_avg(data, stores, c, period)
            cat_vol[c] = total
        return sorted(cats, key=lambda x: cat_vol[x], reverse=True)

    period_names = {"7天": "近7天品退率汇总", "30天": "近30天品退率汇总"}

    # ---- Data sheets ----
    for period in selected_periods:
        sheet_name = period_names[period]
        sorted_cats = sort_cats(all_cats[period], period)

        if progress_callback:
            progress_callback(f"正在创建{sheet_name}...")

        ws = wb.create_sheet(title=sheet_name)
        n_stores = len(stores)
        total_cols = 2 + n_stores * 2

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1, value=f"各店铺品类品退率横向对比 — {sheet_name}")
        c.font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 6

        # Row 3: Header
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(row=3, column=c_idx)
            cell.fill = HEADER_FILL; cell.border = THIN_BORDER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=3, column=1, value="三级品类")
        ws.cell(row=3, column=2, value="内部加权均值")
        col = 3
        for store in stores:
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
            ws.cell(row=3, column=col, value=store)
            col += 2
        ws.row_dimensions[3].height = 24

        # Row 4: Sub-header
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(row=4, column=c_idx)
            cell.fill = HEADER_FILL if c_idx <= 2 else SUB_HEADER_FILL
            cell.font = HEADER_FONT if c_idx <= 2 else SUB_HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        col = 3
        for _ in stores:
            ws.cell(row=4, column=col, value="品退率")
            ws.cell(row=4, column=col + 1, value="达标判定")
            col += 2
        ws.row_dimensions[4].height = 20

        # Data
        row = 5
        for cat in sorted_cats:
            ws.cell(row=row, column=1, value=cat).font = CAT_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row, column=1).border = THIN_BORDER

            internal_avg, _ = calc_internal_avg(data, stores, cat, period)
            avg_cell = ws.cell(row=row, column=2, value=internal_avg if internal_avg is not None else "-")
            avg_cell.font = AVG_FONT; avg_cell.alignment = Alignment(horizontal="center", vertical="center")
            avg_cell.border = THIN_BORDER; avg_cell.fill = AVG_FILL
            if internal_avg is not None:
                avg_cell.number_format = "0.00%"

            col = 3
            for store in stores:
                sd = data[store].get(period, {}).get(cat, {})
                rate = sd.get("rate"); qty = sd.get("qty", 0)
                industry = sd.get("industry", ""); hs = has_sales(sd)

                if hs:
                    rc = ws.cell(row=row, column=col, value=rate)
                    rc.number_format = "0.00%"
                else:
                    rc = ws.cell(row=row, column=col, value="-")
                rc.font = NA_FONT if not hs else DATA_FONT
                rc.alignment = Alignment(horizontal="center", vertical="center")
                rc.border = THIN_BORDER; rc.fill = industry_fill(industry, hs)

                label = industry_label(industry, hs)
                db = ws.cell(row=row, column=col + 1, value=label)
                db.font = Font(name="微软雅黑", size=9, color=label_color(label))
                db.alignment = Alignment(horizontal="center", vertical="center")
                db.border = THIN_BORDER; db.fill = industry_fill(industry, hs)
                col += 2

            ws.row_dimensions[row].height = 20
            row += 1

        ws.freeze_panes = "C5"
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 14
        for i in range(n_stores):
            ws.column_dimensions[get_column_letter(3 + i * 2)].width = 10
            ws.column_dimensions[get_column_letter(4 + i * 2)].width = 9

    # ---- Summary sheet ----
    if progress_callback:
        progress_callback("正在创建汇总概览...")

    ws = wb.create_sheet(title="汇总概览")
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="品退率横向对比 — 汇总概览").font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:F3")
    ws.cell(row=3, column=1, value="颜色图例（按行业水平判定，仅统计有销售的品类）").font = Font(name="微软雅黑", bold=True, size=11)

    legend = [
        (4, "达标 — 优于行业均值", GREEN_FILL),
        (5, "轻微超标 — 行业均值 1-2 倍", YELLOW_FILL),
        (6, "超标 — 行业均值 2-3 倍", ORANGE_FILL),
        (7, "严重超标 — 行业均值 3-5 倍 / 5倍以上", RED_FILL),
        (8, "无销售 — 该店铺未经营此品类", NA_FILL),
    ]
    for r, label, fill in legend:
        c = ws.cell(row=r, column=1, value=f"  {label}")
        c.fill = fill; c.font = DATA_FONT; c.border = THIN_BORDER

    # 严重超标清单
    def write_problem_list(ws, start_row, period, title):
        r = start_row
        ws.merge_cells(f"A{r}:G{r}")
        ws.cell(row=r, column=1, value=title).font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
        r += 1
        headers = ["店铺", "三级品类", "品退率", "行业判定", "订单量", "占店铺订单比", "内部均值参考"]
        for i, h in enumerate(headers):
            c = ws.cell(row=r, column=i + 1, value=h)
            c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = THIN_BORDER
        r += 1
        for store in stores:
            if period not in data[store]:
                continue
            sd = data[store][period]
            store_qty = sum(d["qty"] for d in sd.values())
            problem = []
            for cat, d in sd.items():
                if not has_sales(d): continue
                ind = str(d.get("industry", ""))
                if "3-5倍" in ind or "5倍以上" in ind:
                    problem.append((cat, d["rate"], ind, d["qty"]))
            problem.sort(key=lambda x: x[3], reverse=True)
            first = True
            if problem:
                for cat, rate, ind, qty in problem:
                    ws.cell(row=r, column=1, value=store if first else "").font = DATA_FONT
                    ws.cell(row=r, column=1).border = THIN_BORDER
                    ws.cell(row=r, column=2, value=cat).font = DATA_FONT; ws.cell(row=r, column=2).border = THIN_BORDER
                    rc = ws.cell(row=r, column=3, value=round(rate, 4) if rate else "-")
                    rc.font = DATA_FONT; rc.border = THIN_BORDER; rc.fill = RED_FILL
                    rc.number_format = "0.00%"; rc.alignment = Alignment(horizontal="center")
                    ws.cell(row=r, column=4, value=ind).font = DATA_FONT
                    ws.cell(row=r, column=4).border = THIN_BORDER; ws.cell(row=r, column=4).fill = RED_FILL
                    ws.cell(row=r, column=5, value=qty).font = DATA_FONT
                    ws.cell(row=r, column=5).border = THIN_BORDER; ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
                    pct = f"{qty / store_qty * 100:.1f}%" if store_qty > 0 else "-"
                    ws.cell(row=r, column=6, value=pct).font = DATA_FONT
                    ws.cell(row=r, column=6).border = THIN_BORDER; ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
                    int_avg, _ = calc_internal_avg(data, stores, cat, period)
                    ref = f"{int_avg:.2%}" if int_avg else "-"
                    ws.cell(row=r, column=7, value=ref).font = Font(name="微软雅黑", size=9, color="2F5496")
                    ws.cell(row=r, column=7).border = THIN_BORDER; ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")
                    first = False; r += 1
            else:
                ws.cell(row=r, column=1, value=store).font = DATA_FONT; ws.cell(row=r, column=1).border = THIN_BORDER
                gc = ws.cell(row=r, column=2, value="无严重超标品类")
                gc.font = Font(name="微软雅黑", size=10, color="008000"); gc.fill = GREEN_FILL
                ws.cell(row=r, column=2).border = THIN_BORDER; r += 1
        return r + 1

    # 店铺健康度
    def write_health(ws, start_row, period, title):
        r = start_row
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value=title).font = Font(name="微软雅黑", bold=True, size=12, color="2F5496")
        r += 1
        headers = ["店铺", "经营品类数", "达标", "轻微超标", "超标", "严重超标", "总订单量", "店铺加权品退率"]
        for i, h in enumerate(headers):
            c = ws.cell(row=r, column=i + 1, value=h)
            c.font = Font(name="微软雅黑", bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = THIN_BORDER
        r += 1
        for store in stores:
            if period not in data[store]:
                continue
            sd = data[store][period]
            active = {c: d for c, d in sd.items() if has_sales(d)}
            total = len(active)
            dabiao = sum(1 for d in active.values() if "优于" in str(d.get("industry", "")))
            qingwei = sum(1 for d in active.values() if "1-2倍" in str(d.get("industry", "")))
            chaobiao = sum(1 for d in active.values() if "2-3倍" in str(d.get("industry", "")))
            yanzhong = sum(1 for d in active.values() if "3-5倍" in str(d.get("industry", "")) or "5倍以上" in str(d.get("industry", "")))
            total_qty = sum(d.get("qty", 0) for d in active.values())
            w_sum = 0; w_qty = 0
            for d in active.values():
                if d["rate"] is not None and d["qty"] > 0:
                    w_sum += d["rate"] * d["qty"]; w_qty += d["qty"]
            store_wavg = w_sum / w_qty if w_qty > 0 else None
            vals = [store, total, dabiao, qingwei, chaobiao, yanzhong, total_qty, store_wavg]
            for i, val in enumerate(vals):
                cell = ws.cell(row=r, column=i + 1, value=val)
                cell.font = DATA_FONT; cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER
                if i == 7 and val is not None: cell.number_format = "0.00%"
            ws.cell(row=r, column=3).fill = GREEN_FILL
            if qingwei > 0: ws.cell(row=r, column=4).fill = YELLOW_FILL
            if chaobiao > 0: ws.cell(row=r, column=5).fill = ORANGE_FILL
            if yanzhong > 0: ws.cell(row=r, column=6).fill = RED_FILL
            r += 1
        return r + 1

    r = 10
    for period in selected_periods:
        title = f"严重超标品类清单（行业均值3-5倍及以上）— {period_names[period]}"
        r = write_problem_list(ws, r, period, title)
        r += 1

    for period in selected_periods:
        title = f"店铺健康度总览 — {period_names[period]}（仅统计有销售的品类）"
        r = write_health(ws, r, period, title)
        r += 1

    for col, w in [("A", 22), ("B", 16), ("C", 14), ("D", 16), ("E", 13), ("F", 16), ("G", 12), ("H", 16)]:
        ws.column_dimensions[col].width = w

    if progress_callback:
        progress_callback("正在保存文件...")

    wb.save(output_path)


# ====================== GUI ======================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("品退率横向对比分析工具 v2.0")
        self.root.geometry("820x720")
        self.root.minsize(720, 600)

        self.default_font = ("微软雅黑", 10)
        self.title_font = ("微软雅黑", 14, "bold")
        self.root.option_add("*Font", self.default_font)

        style = ttk.Style()
        style.theme_use("clam")

        self.input_folder = tk.StringVar(value="")
        self.output_folder = tk.StringVar(value="")
        # 周期选择: "both" / "7天" / "30天"
        self.period_choice = tk.StringVar(value="both")

        cfg = load_config()
        if "input_folder" in cfg:
            self.input_folder.set(cfg["input_folder"])
        if "output_folder" in cfg:
            self.output_folder.set(cfg["output_folder"])
        if "period_choice" in cfg:
            self.period_choice.set(cfg["period_choice"])

        self.build_ui()

    def build_ui(self):
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="品退率横向对比分析工具", font=self.title_font).pack(pady=(0, 15))

        # ---- 文件夹选择 ----
        folder_frame = ttk.LabelFrame(main_frame, text="文件路径设置", padding=10)
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(folder_frame, text="源文件目录:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(folder_frame, textvariable=self.input_folder, width=60).grid(row=0, column=1, sticky=tk.EW, padx=5, pady=5)
        ttk.Button(folder_frame, text="选择...", command=self.select_input).grid(row=0, column=2, pady=5)

        ttk.Label(folder_frame, text="输出目录:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(folder_frame, textvariable=self.output_folder, width=60).grid(row=1, column=1, sticky=tk.EW, padx=5, pady=5)
        ttk.Button(folder_frame, text="选择...", command=self.select_output).grid(row=1, column=2, pady=5)

        folder_frame.columnconfigure(1, weight=1)

        # ---- 周期选择 ----
        period_frame = ttk.LabelFrame(main_frame, text="分析周期", padding=10)
        period_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Radiobutton(period_frame, text="7天 + 30天（都要）", variable=self.period_choice, value="both",
                        command=self.on_period_change).grid(row=0, column=0, padx=(0, 20))
        ttk.Radiobutton(period_frame, text="仅 7天", variable=self.period_choice, value="7天",
                        command=self.on_period_change).grid(row=0, column=1, padx=(0, 20))
        ttk.Radiobutton(period_frame, text="仅 30天", variable=self.period_choice, value="30天",
                        command=self.on_period_change).grid(row=0, column=2)

        # ---- 操作按钮 ----
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.scan_btn = ttk.Button(btn_frame, text="1. 扫描文件", command=self.scan_files)
        self.scan_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.generate_btn = ttk.Button(btn_frame, text="2. 生成汇总表", command=self.generate, state=tk.DISABLED)
        self.generate_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.progress = ttk.Progressbar(btn_frame, mode="indeterminate")
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        # ---- 文件匹配结果 ----
        result_frame = ttk.LabelFrame(main_frame, text="自动识别结果（根据文件名自动提取店铺名）", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        columns = ("店铺", "7天文件", "30天文件", "状态")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=8)
        for col in columns:
            self.tree.heading(col, text=col)
        self.tree.column("店铺", width=200)
        self.tree.column("7天文件", width=230)
        self.tree.column("30天文件", width=230)
        self.tree.column("状态", width=80)

        tree_scroll = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # ---- 日志 ----
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=5, wrap=tk.WORD, font=("Consolas", 9))
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # ---- 底部 ----
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Label(bottom_frame, text="自动识别：从文件名中提取店铺名，无需手动配置", foreground="gray").pack(side=tk.LEFT)
        self.status_label = ttk.Label(bottom_frame, text="就绪", foreground="gray")
        self.status_label.pack(side=tk.RIGHT)

    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def set_status(self, msg, color="gray"):
        self.status_label.config(text=msg, foreground=color)

    def select_input(self):
        path = filedialog.askdirectory(title="选择包含各店铺品退率Excel的文件夹")
        if path:
            self.input_folder.set(path)
            self.save_config()

    def select_output(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.output_folder.set(path)
            self.save_config()

    def on_period_change(self):
        self.save_config()

    def save_config(self):
        cfg = {
            "input_folder": self.input_folder.get(),
            "output_folder": self.output_folder.get(),
            "period_choice": self.period_choice.get(),
        }
        save_config(cfg)

    def scan_files(self):
        input_path = self.input_folder.get()
        if not input_path or not os.path.isdir(input_path):
            messagebox.showerror("错误", "请先选择有效的源文件目录")
            return

        self.progress.start()
        self.set_status("正在扫描...")
        self.log("========== 开始扫描 ==========")

        self.matches, unmatched = scan_folder(input_path)

        # Update tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        all_periods = set()
        for store in sorted(self.matches.keys()):
            files_7 = self.matches[store].get("7天", "")
            files_30 = self.matches[store].get("30天", "")
            if files_7:
                files_7 = os.path.basename(files_7)
                all_periods.add("7天")
            if files_30:
                files_30 = os.path.basename(files_30)
                all_periods.add("30天")

            if files_7 and files_30:
                status = "完整"
            elif files_7:
                status = "缺30天"
            elif files_30:
                status = "缺7天"
            else:
                status = "无数据"

            self.tree.insert("", tk.END, values=(store, files_7 or "未找到", files_30 or "未找到", status))

        if unmatched:
            self.log(f"[警告] {len(unmatched)} 个文件无法识别周期：")
            for f in unmatched:
                self.log(f"  - {f}")

        self.log(f"识别到 {len(self.matches)} 家店铺")
        for p in sorted(all_periods):
            count = sum(1 for s in self.matches if p in self.matches[s])
            self.log(f"  - 有{p}数据的: {count} 家")

        # Auto-adjust period choice if some period is missing
        if "7天" not in all_periods and "30天" in all_periods:
            self.log("[提示] 未检测到7天数据，建议选择「仅30天」")
        elif "30天" not in all_periods and "7天" in all_periods:
            self.log("[提示] 未检测到30天数据，建议选择「仅7天」")

        has_data = any(len(self.matches[s]) >= 1 for s in self.matches)
        if has_data:
            self.generate_btn.config(state=tk.NORMAL)
            self.set_status(f"扫描完成，{len(self.matches)}家店铺", "green")
        else:
            self.generate_btn.config(state=tk.DISABLED)
            self.set_status("扫描完成，无有效数据", "red")

        self.progress.stop()
        self.save_config()

    def generate(self):
        input_path = self.input_folder.get()
        output_path = self.output_folder.get()
        period_choice = self.period_choice.get()

        if not input_path or not os.path.isdir(input_path):
            messagebox.showerror("错误", "请先选择源文件目录")
            return
        if not output_path or not os.path.isdir(output_path):
            messagebox.showerror("错误", "请先选择输出目录")
            return

        # Re-scan
        self.matches, unmatched = scan_folder(input_path)

        # Determine which periods to generate
        available_periods = set()
        for store in self.matches:
            available_periods.update(self.matches[store].keys())

        if period_choice == "both":
            selected = []
            if "7天" in available_periods:
                selected.append("7天")
            if "30天" in available_periods:
                selected.append("30天")
            # Also build stores list from both
            stores = sorted(self.matches.keys())
        elif period_choice == "7天":
            selected = ["7天"] if "7天" in available_periods else []
            stores = sorted([s for s in self.matches if "7天" in self.matches[s]])
        else:  # "30天"
            selected = ["30天"] if "30天" in available_periods else []
            stores = sorted([s for s in self.matches if "30天" in self.matches[s]])

        if not selected:
            messagebox.showerror("错误", f"没有找到{period_choice}的数据")
            return

        if not stores:
            messagebox.showerror("错误", "没有找到任何有效数据")
            return

        if period_choice == "both":
            suffix = "近7天_近30天"
        elif period_choice == "7天":
            suffix = "近7天"
        else:
            suffix = "近30天"

        output_file = os.path.join(output_path, f"品退率横向对比汇总_{suffix}.xlsx")

        self.progress.start()
        self.generate_btn.config(state=tk.DISABLED)

        def run():
            try:
                data = load_all_data(self.matches)
                def cb(msg):
                    self.log(f"[进度] {msg}")
                    self.set_status(msg)

                self.log("========== 开始生成 ==========")
                self.log(f"涉及店铺: {', '.join(stores)}")
                self.log(f"分析周期: {', '.join(selected)}")

                generate_excel(data, stores, output_file, selected, progress_callback=cb)

                sheet_names = {"7天": "近7天品退率汇总", "30天": "近30天品退率汇总"}
                self.log(f"生成的Sheet: 汇总概览 + {', '.join(sheet_names[p] for p in selected)}")
                self.log(f"✅ 生成完成: {output_file}")
                self.set_status("生成完成!", "green")
                self.root.after(0, lambda: messagebox.showinfo("完成", f"汇总表已生成:\n{output_file}"))
            except Exception as e:
                import traceback
                self.log(f"[错误] {traceback.format_exc()}")
                self.set_status("生成失败", "red")
                self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
            finally:
                self.progress.stop()
                self.generate_btn.config(state=tk.NORMAL)

        threading.Thread(target=run, daemon=True).start()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
